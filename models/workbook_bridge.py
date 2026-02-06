"""Bridge between the audited V8 workbook and the Python EPM engine."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook


YEAR_COLS = {2026: 2, 2027: 4, 2028: 6, 2029: 8, 2030: 10}
YEARS = [2026, 2027, 2028, 2029, 2030]


def _year_months(year: int) -> List[str]:
    return [f"{year}-{month:02d}" for month in range(1, 13)]


@dataclass
class WorkbookBaseline:
    revenue: Dict[int, float] = field(default_factory=dict)
    cogs: Dict[int, float] = field(default_factory=dict)
    opex: Dict[int, float] = field(default_factory=dict)
    ebitda: Dict[int, float] = field(default_factory=dict)
    delta_wc: Dict[int, float] = field(default_factory=dict)
    capex: Dict[int, float] = field(default_factory=dict)
    ending_cash: Dict[int, float] = field(default_factory=dict)
    enterprise_value: float = 0.0


def _year_values(ws, row: int) -> Dict[int, float]:
    values = {}
    for year, col in YEAR_COLS.items():
        val = ws.cell(row, col).value
        values[year] = float(val or 0.0)
    return values


def _safe_positive(val: float) -> float:
    return val if val > 0 else 0.0


def compute_workbook_baseline(workbook_path: Path) -> WorkbookBaseline:
    """
    Compute workbook-equivalent annual outputs from input rows.

    This avoids Excel formula evaluation while matching the linked-sheet logic
    used in the V8 audited model.
    """
    wb = load_workbook(workbook_path, data_only=False)
    inputs = wb["00_Inputs_Assumptions"]
    opex_sheet = wb["15_OpEx_Detail"]
    capex_sheet = wb["16_CapEx_Detail"]
    bom_sheet = wb["14_Unit_Costs_BOM"]
    val_sheet = wb["19_Valuation_Models"]

    feed_per_animal = float(inputs.cell(6, 2).value or 0.0)
    litter_per_animal = float(inputs.cell(7, 2).value or 0.0)
    cash_buffer = float(inputs.cell(5, 2).value or 0.0)
    dio = float(inputs.cell(9, 2).value or 0.0)
    dpo = float(inputs.cell(10, 2).value or 0.0)
    tax_rate = float(inputs.cell(11, 2).value or 0.0)

    # Segment inputs.
    direct_active = _year_values(inputs, 19)
    direct_animals = _year_values(inputs, 20)
    direct_inclusion = _year_values(inputs, 21)
    direct_fee = _year_values(inputs, 23)
    direct_base = _year_values(inputs, 42)
    direct_premium = _year_values(inputs, 43)
    direct_wtp = _year_values(inputs, 44)

    partner_signed = _year_values(inputs, 26)
    partner_animals = _year_values(inputs, 27)
    partner_inclusion = _year_values(inputs, 28)
    partner_fee = _year_values(inputs, 30)
    partner_base = _year_values(inputs, 45)
    partner_premium = _year_values(inputs, 46)
    partner_wtp = _year_values(inputs, 47)
    partner_conv = _year_values(inputs, 51)
    partner_churn = _year_values(inputs, 52)

    litter_active = _year_values(inputs, 33)
    litter_animals = _year_values(inputs, 34)
    litter_inclusion = _year_values(inputs, 35)
    litter_fee = _year_values(inputs, 37)
    litter_base = _year_values(inputs, 48)
    litter_premium = _year_values(inputs, 49)
    litter_wtp = _year_values(inputs, 50)

    dso_direct = _year_values(inputs, 55)
    dso_partner = _year_values(inputs, 56)
    dso_litter = _year_values(inputs, 57)
    unit_factor = _year_values(inputs, 40)

    equity_seed = _year_values(inputs, 13)
    equity_pre = _year_values(inputs, 14)
    equity_series = _year_values(inputs, 15)

    # Opex fixed block.
    fixed_opex = {year: 0.0 for year in YEARS}
    for row in range(5, 20):
        row_vals = _year_values(opex_sheet, row)
        for year in YEARS:
            fixed_opex[year] += row_vals[year]

    # Capex and depreciation blocks.
    capex_lab = _year_values(capex_sheet, 5)
    capex_pilot = _year_values(capex_sheet, 6)
    capex_valdarno = _year_values(capex_sheet, 7)
    capex_it = _year_values(capex_sheet, 8)
    total_capex = {
        y: capex_lab[y] + capex_pilot[y] + capex_valdarno[y] + capex_it[y] for y in YEARS
    }

    # Unit production costs from BOM tab formulas.
    raw_total = 0.0
    for row in list(range(6, 16)) + list(range(19, 26)):
        pct = float(bom_sheet.cell(row, 2).value or 0.0)
        eurkg = float(bom_sheet.cell(row, 3).value or 0.0)
        raw_total += pct * eurkg
    industrial_total = sum(float(bom_sheet.cell(r, 4).value or 0.0) for r in range(31, 35))
    chitosano = float(bom_sheet.cell(9, 2).value or 0.0) * float(bom_sheet.cell(9, 3).value or 0.0)
    unit_cost_poultry_base = raw_total + industrial_total
    unit_cost_litter_base = unit_cost_poultry_base - chitosano

    # Partner active customers are recursive in workbook.
    partner_active = {}
    prev_year = YEARS[0]
    partner_active[prev_year] = _safe_positive(partner_signed[prev_year] * partner_conv[prev_year])
    for year in YEARS[1:]:
        partner_active[year] = _safe_positive(
            partner_active[prev_year] * (1.0 - partner_churn[year])
            + partner_signed[prev_year] * partner_conv[year]
        )
        prev_year = year

    # Segment annual quantities and revenues.
    direct_units = {}
    partner_units = {}
    litter_units = {}
    direct_revenue = {}
    partner_revenue = {}
    litter_revenue = {}
    total_revenue = {}
    total_cogs = {}
    gross_profit = {}

    for year in YEARS:
        price_direct = max(0.0, direct_base[year] + direct_premium[year] * direct_wtp[year])
        price_partner = max(0.0, partner_base[year] + partner_premium[year] * partner_wtp[year])
        price_litter = max(0.0, litter_base[year] + litter_premium[year] * litter_wtp[year])

        direct_units[year] = direct_active[year] * direct_animals[year] * feed_per_animal * direct_inclusion[year]
        partner_units[year] = partner_active[year] * partner_animals[year] * feed_per_animal * partner_inclusion[year]
        litter_units[year] = litter_active[year] * litter_animals[year] * litter_per_animal * litter_inclusion[year]

        direct_revenue[year] = direct_units[year] * price_direct + direct_active[year] * direct_fee[year]
        partner_revenue[year] = partner_units[year] * price_partner + partner_active[year] * partner_fee[year]
        litter_revenue[year] = litter_units[year] * price_litter + litter_active[year] * litter_fee[year]
        total_revenue[year] = direct_revenue[year] + partner_revenue[year] + litter_revenue[year]

        poultry_unit_cost = unit_cost_poultry_base * unit_factor[year]
        litter_unit_cost = unit_cost_litter_base * unit_factor[year]
        total_cogs[year] = (
            (direct_units[year] + partner_units[year]) * poultry_unit_cost
            + litter_units[year] * litter_unit_cost
        )
        gross_profit[year] = total_revenue[year] - total_cogs[year]

    # Working capital and deltas.
    ar = {}
    inventory = {}
    ap = {}
    net_wc = {}
    delta_wc = {}
    prev_nwc = 0.0
    for year in YEARS:
        ar[year] = (
            direct_revenue[year] * dso_direct[year]
            + partner_revenue[year] * dso_partner[year]
            + litter_revenue[year] * dso_litter[year]
        ) / 365.0
        inventory[year] = total_cogs[year] * dio / 365.0
        ap[year] = total_cogs[year] * dpo / 365.0
        net_wc[year] = ar[year] + inventory[year] - ap[year]
        delta_wc[year] = net_wc[year] - prev_nwc
        prev_nwc = net_wc[year]

    # Depreciation logic mirrors CapEx tab.
    cumul_lab = cumul_pilot = cumul_valdarno = cumul_it = 0.0
    depreciation = {}
    for year in YEARS:
        cumul_lab += capex_lab[year]
        cumul_pilot += capex_pilot[year]
        cumul_valdarno += capex_valdarno[year]
        cumul_it += capex_it[year]
        depreciation[year] = (
            cumul_lab / 5.0
            + cumul_pilot / 7.0
            + cumul_valdarno / 10.0
            + cumul_it / 3.0
        )

    # Cash loop with contingency formula.
    limits = {2026: 20000.0, 2027: 30000.0, 2028: 40000.0, 2029: 50000.0, 2030: 60000.0}
    total_opex = {}
    ebitda = {}
    taxes = {}
    ending_cash = {}
    unlevered_fcf = {}
    beginning_cash = 0.0

    for year in YEARS:
        equity_inflow = equity_seed[year] + equity_pre[year] + equity_series[year]
        contingency = min(
            limits[year],
            max(
                0.0,
                beginning_cash
                + equity_inflow
                + gross_profit[year]
                - fixed_opex[year]
                - delta_wc[year]
                - total_capex[year]
                - cash_buffer,
            ),
        )
        total_opex[year] = fixed_opex[year] + contingency
        ebitda[year] = gross_profit[year] - total_opex[year]

        ebit = ebitda[year] - depreciation[year]
        taxes[year] = -ebit * tax_rate if ebit > 0 else 0.0

        net_change = equity_inflow + ebitda[year] + taxes[year] - delta_wc[year] - total_capex[year]
        ending_cash[year] = beginning_cash + net_change
        unlevered_fcf[year] = ebitda[year] + taxes[year] - delta_wc[year] - total_capex[year]
        beginning_cash = ending_cash[year]

    # DCF block.
    discount_rate = float(val_sheet.cell(5, 2).value or 0.25)
    terminal_growth = float(val_sheet.cell(6, 2).value or 0.03)
    discount_factors = {year: 1.0 / ((1.0 + discount_rate) ** (idx + 1)) for idx, year in enumerate(YEARS)}
    pv_fcf = {year: unlevered_fcf[year] * discount_factors[year] for year in YEARS}
    terminal_value = unlevered_fcf[2030] * (1.0 + terminal_growth) / (discount_rate - terminal_growth)
    pv_terminal = terminal_value * discount_factors[2030]
    enterprise_value = sum(pv_fcf.values()) + pv_terminal

    return WorkbookBaseline(
        revenue=total_revenue,
        cogs=total_cogs,
        opex=total_opex,
        ebitda=ebitda,
        delta_wc=delta_wc,
        capex=total_capex,
        ending_cash=ending_cash,
        enterprise_value=enterprise_value,
    )


def build_assumptions_from_workbook(workbook_path: Path) -> dict:
    """
    Create an EPM assumptions object derived from workbook V8 inputs.

    The Python model is monthly and deterministic, so annual workbook figures
    are translated into monthly step inputs.
    """
    wb = load_workbook(workbook_path, data_only=False)
    inputs = wb["00_Inputs_Assumptions"]
    baseline = compute_workbook_baseline(workbook_path)

    feed_per_animal = float(inputs.cell(6, 2).value or 0.0)
    litter_per_animal = float(inputs.cell(7, 2).value or 0.0)
    dio = float(inputs.cell(9, 2).value or 0.0)
    dpo = float(inputs.cell(10, 2).value or 0.0)

    # Segment values used to derive unit share and blended segment prices.
    direct_active = _year_values(inputs, 19)
    direct_animals = _year_values(inputs, 20)
    direct_inclusion = _year_values(inputs, 21)
    direct_fee = _year_values(inputs, 23)
    direct_base = _year_values(inputs, 42)
    direct_premium = _year_values(inputs, 43)
    direct_wtp = _year_values(inputs, 44)

    partner_signed = _year_values(inputs, 26)
    partner_animals = _year_values(inputs, 27)
    partner_inclusion = _year_values(inputs, 28)
    partner_fee = _year_values(inputs, 30)
    partner_base = _year_values(inputs, 45)
    partner_premium = _year_values(inputs, 46)
    partner_wtp = _year_values(inputs, 47)
    partner_conv = _year_values(inputs, 51)
    partner_churn = _year_values(inputs, 52)

    litter_active = _year_values(inputs, 33)
    litter_animals = _year_values(inputs, 34)
    litter_inclusion = _year_values(inputs, 35)
    litter_fee = _year_values(inputs, 37)
    litter_base = _year_values(inputs, 48)
    litter_premium = _year_values(inputs, 49)
    litter_wtp = _year_values(inputs, 50)

    dso_direct = _year_values(inputs, 55)
    dso_partner = _year_values(inputs, 56)
    dso_litter = _year_values(inputs, 57)

    partner_active = {}
    prev_year = YEARS[0]
    partner_active[prev_year] = _safe_positive(partner_signed[prev_year] * partner_conv[prev_year])
    for year in YEARS[1:]:
        partner_active[year] = _safe_positive(
            partner_active[prev_year] * (1.0 - partner_churn[year])
            + partner_signed[prev_year] * partner_conv[year]
        )
        prev_year = year

    units_direct = {}
    units_partner = {}
    units_litter = {}
    prices_direct = {}
    prices_partner = {}
    prices_litter = {}
    dso_weighted = {}

    for year in YEARS:
        base_price_direct = max(0.0, direct_base[year] + direct_premium[year] * direct_wtp[year])
        base_price_partner = max(0.0, partner_base[year] + partner_premium[year] * partner_wtp[year])
        base_price_litter = max(0.0, litter_base[year] + litter_premium[year] * litter_wtp[year])

        units_direct[year] = direct_active[year] * direct_animals[year] * feed_per_animal * direct_inclusion[year]
        units_partner[year] = partner_active[year] * partner_animals[year] * feed_per_animal * partner_inclusion[year]
        units_litter[year] = litter_active[year] * litter_animals[year] * litter_per_animal * litter_inclusion[year]
        total_units = units_direct[year] + units_partner[year] + units_litter[year]

        fee_component_direct = (direct_active[year] * direct_fee[year] / units_direct[year]) if units_direct[year] > 0 else 0.0
        fee_component_partner = (partner_active[year] * partner_fee[year] / units_partner[year]) if units_partner[year] > 0 else 0.0
        fee_component_litter = (litter_active[year] * litter_fee[year] / units_litter[year]) if units_litter[year] > 0 else 0.0

        prices_direct[year] = base_price_direct + fee_component_direct
        prices_partner[year] = base_price_partner + fee_component_partner
        prices_litter[year] = base_price_litter + fee_component_litter

        if baseline.revenue[year] > 0:
            direct_share = (units_direct[year] * prices_direct[year]) / baseline.revenue[year]
            partner_share = (units_partner[year] * prices_partner[year]) / baseline.revenue[year]
            litter_share = (units_litter[year] * prices_litter[year]) / baseline.revenue[year]
            dso_weighted[year] = (
                direct_share * dso_direct[year]
                + partner_share * dso_partner[year]
                + litter_share * dso_litter[year]
            )
        else:
            dso_weighted[year] = dso_direct[year]

    # Build monthly step values.
    capacity_by_month = {}
    direct_price_by_month = {}
    partner_price_by_month = {}
    litter_price_by_month = {}
    direct_mix = {}
    partner_mix = {}
    litter_mix = {}
    capex_by_month = {}
    equity_by_month = {}
    cogs_direct_price = {}
    cogs_partner_price = {}
    cogs_litter_price = {}
    total_opex_monthly = {}

    for year in YEARS:
        jan = f"{year}-01"
        total_units = units_direct[year] + units_partner[year] + units_litter[year]
        for month in _year_months(year):
            capacity_by_month[month] = total_units / 12.0
        direct_price_by_month[jan] = prices_direct[year]
        partner_price_by_month[jan] = prices_partner[year]
        litter_price_by_month[jan] = prices_litter[year]
        direct_mix[str(year)] = units_direct[year] / total_units if total_units > 0 else 0.0
        partner_mix[str(year)] = units_partner[year] / total_units if total_units > 0 else 0.0
        litter_mix[str(year)] = units_litter[year] / total_units if total_units > 0 else 0.0
        capex_by_month[jan] = baseline.capex[year]
        equity_by_month[jan] = (
            float(inputs.cell(13, YEAR_COLS[year]).value or 0.0)
            + float(inputs.cell(14, YEAR_COLS[year]).value or 0.0)
            + float(inputs.cell(15, YEAR_COLS[year]).value or 0.0)
        )

        direct_units_value = units_direct[year]
        partner_units_value = units_partner[year]
        litter_units_value = units_litter[year]
        cogs_direct_price[jan] = (
            (baseline.cogs[year] * (direct_units_value / (direct_units_value + partner_units_value + litter_units_value)))
            / direct_units_value
            if direct_units_value > 0
            else 0.0
        )
        cogs_partner_price[jan] = (
            (baseline.cogs[year] * (partner_units_value / (direct_units_value + partner_units_value + litter_units_value)))
            / partner_units_value
            if partner_units_value > 0
            else 0.0
        )
        cogs_litter_price[jan] = (
            (baseline.cogs[year] * (litter_units_value / (direct_units_value + partner_units_value + litter_units_value)))
            / litter_units_value
            if litter_units_value > 0
            else 0.0
        )
        for month in _year_months(year):
            total_opex_monthly[month] = baseline.opex[year] / 12.0

    avg_dso = int(round(sum(dso_weighted.values()) / len(dso_weighted)))

    opex_base_monthly = total_opex_monthly.get("2026-01", 0.0) or 1.0
    opex_ramp = {
        f"{year}-01": (total_opex_monthly.get(f"{year}-01", opex_base_monthly) / opex_base_monthly)
        for year in YEARS
    }

    assumptions = {
        "meta": {
            "project": "ReSemis EPM - Workbook V8 Bridge",
            "source_workbook": str(workbook_path),
            "granularity": "monthly",
            "currency": "EUR",
            "unit_base": "kg",
        },
        "time_horizon": {"start_month": "2026-01", "end_month": "2030-12"},
        "products": [
            {"product_id": "poultry_direct", "product_name": "Poultry Direct", "unit": "kg"},
            {"product_id": "poultry_partner", "product_name": "Poultry Partner", "unit": "kg"},
            {"product_id": "litter_soil", "product_name": "Litter Soil B2B2B", "unit": "kg"},
        ],
        "markets": [{"market_id": "global", "geo": "EU", "activation_month": "2026-01"}],
        "scenarios": [
            {"scenario_id": "base", "description": "Workbook V8 base bridge"},
            {"scenario_id": "conservative", "description": "Downside bridge"},
            {"scenario_id": "aggressive", "description": "Upside bridge"},
        ],
        "volume": {
            "tam": {"per_market_kg": {"global": 1e12}},
            "sam_share": {"per_market_pct": {"global": 1.0}},
            "som_share": {
                "per_market_pct": {"global": 1.0},
                "ramp": {"by_market": {"global": {"start_month": "2026-01", "duration_months": 0, "curve": "linear"}}},
            },
            "capacity": {"enabled": True, "by_month": capacity_by_month},
        },
        "mix": {
            "by_market": {
                "global": {
                    "by_product": {
                        "poultry_direct": {"by_year": direct_mix},
                        "poultry_partner": {"by_year": partner_mix},
                        "litter_soil": {"by_year": litter_mix},
                    }
                }
            }
        },
        "pricing": {
            "list_price": {
                "by_product": {
                    "poultry_direct": {"by_month": direct_price_by_month},
                    "poultry_partner": {"by_month": partner_price_by_month},
                    "litter_soil": {"by_month": litter_price_by_month},
                }
            },
            "discounts": {
                "by_product": {
                    "poultry_direct": {"by_market": {"global": {"by_month": {}}}},
                    "poultry_partner": {"by_market": {"global": {"by_month": {}}}},
                    "litter_soil": {"by_market": {"global": {"by_month": {}}}},
                }
            },
        },
        "bom": {
            "by_product": {
                "poultry_direct": {"inputs": [{"input_id": "rm_direct", "input_name": "Direct RM", "qty_per_kg": 1.0}]},
                "poultry_partner": {"inputs": [{"input_id": "rm_partner", "input_name": "Partner RM", "qty_per_kg": 1.0}]},
                "litter_soil": {"inputs": [{"input_id": "rm_litter", "input_name": "Litter RM", "qty_per_kg": 1.0}]},
            }
        },
        "input_prices": {
            "by_input": {
                "rm_direct": {"base_price": 0.0, "by_month": cogs_direct_price},
                "rm_partner": {"base_price": 0.0, "by_month": cogs_partner_price},
                "rm_litter": {"base_price": 0.0, "by_month": cogs_litter_price},
            }
        },
        "fixed_cogs": {"base_monthly": 0.0, "ramp": {"by_month": {}}},
        "opex": {
            "fixed": {"by_category": {"management_ga": {"base_monthly": opex_base_monthly, "ramp": {"by_month": opex_ramp}}}},
            "variable": {"by_driver": {}},
            "sales_marketing": {"fixed_base": 0.0, "ramp": {"by_month": {}}, "cac": {"by_market": {"global": 0.0}}},
        },
        "working_capital": {"dso_days": avg_dso, "dio_days": int(dio), "dpo_days": int(dpo)},
        "capex": {"base_monthly": 0.0, "milestones": {"by_month": capex_by_month}},
        "funding": {
            "initial_cash": 0.0,
            "equity": {"by_month": equity_by_month},
            "debt": {"interest_rate": 0.0, "by_month": {}},
        },
        "valuation": {
            "discount_rate": float(wb["19_Valuation_Models"].cell(5, 2).value or 0.25),
            "terminal_growth_rate": float(wb["19_Valuation_Models"].cell(6, 2).value or 0.03),
            "terminal_method": "gordon",
            "terminal_multiple": 3.0,
            "exit_year": 2030,
            "equity": {"ownership_pct": 1.0, "invested": {"by_month": equity_by_month}},
        },
    }

    return assumptions


def annualize_series(monthly: Dict[str, float], mode: str = "sum") -> Dict[int, float]:
    """Aggregate monthly series into annual values."""
    annual: Dict[int, float] = {year: 0.0 for year in YEARS}
    if mode == "end":
        for month, value in sorted(monthly.items()):
            year = int(month[:4])
            annual[year] = value
        return annual

    for month, value in monthly.items():
        year = int(month[:4])
        annual[year] = annual.get(year, 0.0) + value
    return annual


def reconcile_engine_to_workbook(
    workbook_path: Path,
    revenue_total_monthly: Dict[str, float],
    cogs_total_monthly: Dict[str, float],
    opex_total_monthly: Dict[str, float],
    ebitda_monthly: Dict[str, float],
    ending_cash_monthly: Dict[str, float],
    enterprise_value: float,
) -> Dict[str, Dict[str, float]]:
    """Return annual and scalar variance metrics between engine and workbook baseline."""
    baseline = compute_workbook_baseline(workbook_path)
    engine_revenue = annualize_series(revenue_total_monthly, mode="sum")
    engine_cogs = annualize_series(cogs_total_monthly, mode="sum")
    engine_opex = annualize_series(opex_total_monthly, mode="sum")
    engine_ebitda = annualize_series(ebitda_monthly, mode="sum")
    engine_cash = annualize_series(ending_cash_monthly, mode="end")

    def _variance(engine_val: float, base_val: float) -> float:
        if abs(base_val) < 1e-9:
            return 0.0 if abs(engine_val) < 1e-9 else 1.0
        return (engine_val - base_val) / abs(base_val)

    variances = {"revenue": {}, "cogs": {}, "opex": {}, "ebitda": {}, "ending_cash": {}}
    for year in YEARS:
        variances["revenue"][str(year)] = _variance(engine_revenue[year], baseline.revenue[year])
        variances["cogs"][str(year)] = _variance(engine_cogs[year], baseline.cogs[year])
        variances["opex"][str(year)] = _variance(engine_opex[year], baseline.opex[year])
        variances["ebitda"][str(year)] = _variance(engine_ebitda[year], baseline.ebitda[year])
        variances["ending_cash"][str(year)] = _variance(engine_cash[year], baseline.ending_cash[year])
    variances["enterprise_value"] = {"value": _variance(enterprise_value, baseline.enterprise_value)}

    return variances
